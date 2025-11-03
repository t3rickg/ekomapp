#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import time
import json
import pickle
from datetime import datetime
from colorama import Fore, Back, Style, init

# Renkleri başlat
init(autoreset=True)

class EkomCode:
    def __init__(self):
        self.user_data = {}
        self.current_level = "başlangıç"
        self.progress_file = "user_progress.pkl"
        self.load_user_progress()
        
        # Eğitim içeriği
        self.modules = {
            "python_temelleri": {
                "title": "Python Temelleri",
                "lessons": {
                    "degiskenler": "Değişkenler ve Veri Tipleri",
                    "operatorler": "Operatörler",
                    "kosul_ifadeleri": "Koşul İfadeleri (if-elif-else)",
                    "donguler": "Döngüler (for, while)",
                    "fonksiyonlar": "Fonksiyonlar"
                }
            },
            "otomasyon_egitim": {
                "title": "Otomasyon Projeleri",
                "lessons": {
                    "dosya_okuma": "Dosya Okuma/Yazma Otomasyonu",
                    "web_otomasyon": "Web Otomasyonu",
                    "excel_otomasyon": "Excel Otomasyonu",
                    "mail_otomasyon": "E-posta Otomasyonu",
                    "veri_cekme": "Web'den Veri Çekme"
                }
            },
            "ornek_projeler": {
                "title": "Örnek Projeler",
                "lessons": {
                    "hesap_makinesi": "Hesap Makinesi",
                    "todo_app": "Yapılacaklar Listesi",
                    "password_generator": "Şifre Üretici",
                    "web_scraper": "Web Kazıyıcı",
                    "file_organizer": "Dosya Organizatörü"
                }
            }
        }

    def clear_screen(self):
        """Ekranı temizle"""
        os.system('cls' if os.name == 'nt' else 'clear')

    def print_header(self):
        """Başlık yazdır"""
        header = f"""
{Fore.CYAN}
╔══════════════════════════════════════════════════════════════╗
║                   {Fore.YELLOW}E K O M C O D E{Fore.CYAN}                           ║
║              Python Eğitim ve Otomasyon Platformu           ║
╚══════════════════════════════════════════════════════════════╝
{Style.RESET_ALL}
        """
        print(header)

    def print_menu(self, title, options):
        """Menü yazdır"""
        print(f"\n{Fore.GREEN}╔══ {title} {Fore.GREEN}══╗")
        for key, option in options.items():
            print(f"║ {Fore.YELLOW}{key}.{Style.RESET_ALL} {option}")
        print(f"╚{'═' * (len(title) + 8)}╝")

    def load_user_progress(self):
        """Kullanıcı ilerlemesini yükle"""
        try:
            with open(self.progress_file, 'rb') as f:
                self.user_data = pickle.load(f)
        except FileNotFoundError:
            self.user_data = {
                "start_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "completed_lessons": [],
                "current_module": "python_temelleri",
                "score": 0
            }

    def save_user_progress(self):
        """Kullanıcı ilerlemesini kaydet"""
        with open(self.progress_file, 'wb') as f:
            pickle.dump(self.user_data, f)

    def show_progress(self):
        """İlerlemeyi göster"""
        completed = len(self.user_data["completed_lessons"])
        total = sum(len(module["lessons"]) for module in self.modules.values())
        progress = (completed / total) * 100 if total > 0 else 0
        
        print(f"\n{Fore.CYAN}╔══════════════ İLERLEME DURUMU ══════════════╗")
        print(f"║ {Fore.GREEN}Tamamlanan Dersler: {completed}/{total}")
        print(f"║ {Fore.BLUE}İlerleme: {progress:.1f}%")
        print(f"║ {Fore.YELLOW}Puan: {self.user_data['score']}")
        print(f"║ {Fore.MAGENTA}Başlangıç: {self.user_data['start_date']}")
        print(f"╚═══════════════════════════════════════════════╝")

    def main_menu(self):
        """Ana menü"""
        while True:
            self.clear_screen()
            self.print_header()
            self.show_progress()
            
            menu_options = {
                "1": "Python Temelleri",
                "2": "Otomasyon Eğitimi", 
                "3": "Örnek Projeler",
                "4": "Kod Örneklerini İncele",
                "5": "Alıştırma Yap",
                "6": "Ayarlar",
                "0": "Çıkış"
            }
            
            self.print_menu("ANA MENÜ", menu_options)
            
            choice = input(f"\n{Fore.CYAN}Seçiminiz (0-6): {Style.RESET_ALL}")
            
            if choice == "1":
                self.python_basics_menu()
            elif choice == "2":
                self.automation_menu()
            elif choice == "3":
                self.projects_menu()
            elif choice == "4":
                self.code_examples()
            elif choice == "5":
                self.practice_exercises()
            elif choice == "6":
                self.settings_menu()
            elif choice == "0":
                print(f"\n{Fore.GREEN}EkomCode'u kullandığınız için teşekkürler! 🚀")
                self.save_user_progress()
                break
            else:
                input(f"{Fore.RED}Geçersiz seçim! Tekrar deneyin. (Enter)")

    def python_basics_menu(self):
        """Python temelleri menüsü"""
        self.module_menu("python_temelleri")

    def automation_menu(self):
        """Otomasyon menüsü"""
        self.module_menu("otomasyon_egitim")

    def projects_menu(self):
        """Projeler menüsü"""
        self.module_menu("ornek_projeler")

    def module_menu(self, module_key):
        """Modül menüsünü göster"""
        module = self.modules[module_key]
        
        while True:
            self.clear_screen()
            print(f"\n{Fore.CYAN}╔══════════ {module['title']} ══════════╗")
            
            lessons = module["lessons"]
            for i, (key, lesson) in enumerate(lessons.items(), 1):
                status = "✓" if key in self.user_data["completed_lessons"] else " "
                print(f"║ {i}. [{status}] {lesson}")
            
            print(f"║")
            print(f"║ 0. Ana Menü")
            print(f"╚══════════════════════════════════╝")
            
            try:
                choice = input(f"\n{Fore.CYAN}Ders seçin (1-{len(lessons)}): {Style.RESET_ALL}")
                
                if choice == "0":
                    break
                
                choice_int = int(choice)
                if 1 <= choice_int <= len(lessons):
                    lesson_key = list(lessons.keys())[choice_int - 1]
                    self.show_lesson(module_key, lesson_key)
                else:
                    input(f"{Fore.RED}Geçersiz seçim! (Enter)")
                    
            except ValueError:
                input(f"{Fore.RED}Lütfen sayı girin! (Enter)")

    def show_lesson(self, module_key, lesson_key):
        """Ders içeriğini göster"""
        self.clear_screen()
        
        lesson_content = self.get_lesson_content(module_key, lesson_key)
        code_example = self.get_code_example(module_key, lesson_key)
        
        print(f"\n{Fore.CYAN}╔══════════ {lesson_content['title']} ══════════╗")
        print(f"║{Style.RESET_ALL}")
        
        # Teori
        for line in lesson_content["theory"]:
            print(f"║ {line}")
        
        print(f"║{Style.RESET_ALL}")
        print(f"║ {Fore.YELLOW}Örnek Kod:{Style.RESET_ALL}")
        print(f"║{Style.RESET_ALL}")
        
        # Kod örneği
        for line in code_example.split('\n'):
            print(f"║ {Fore.GREEN}{line}{Style.RESET_ALL}")
        
        print(f"║{Style.RESET_ALL}")
        print(f"╚══════════════════════════════════════════╝")
        
        # Kodu çalıştırma seçeneği
        if lesson_key not in self.user_data["completed_lessons"]:
            run_code = input(f"\n{Fore.CYAN}Kodu çalıştırmak ister misiniz? (e/h): {Style.RESET_ALL}").lower()
            if run_code == 'e':
                self.run_code_example(code_example)
                self.user_data["completed_lessons"].append(lesson_key)
                self.user_data["score"] += 10
                self.save_user_progress()
                print(f"{Fore.GREEN}✓ Ders tamamlandı! +10 puan")
        
        input(f"\n{Fore.CYAN}Ana menüye dönmek için Enter...{Style.RESET_ALL}")

    def get_lesson_content(self, module_key, lesson_key):
        """Ders içeriğini getir"""
        lessons_content = {
            "python_temelleri": {
                "degiskenler": {
                    "title": "Değişkenler ve Veri Tipleri",
                    "theory": [
                        "Değişkenler: Verileri saklamak için kullanılan etiketler",
                        "Veri Tipleri:",
                        "  - int: Tam sayılar (5, -3, 100)",
                        "  - float: Ondalıklı sayılar (3.14, -0.5)",
                        "  - str: Metinler ('Merhaba', \"Python\")",
                        "  - bool: Mantıksal değerler (True, False)",
                        "  - list: Liste [1, 2, 3]",
                        "  - dict: Sözlük {'isim': 'Ali', 'yas': 25}"
                    ]
                },
                "operatorler": {
                    "title": "Operatörler",
                    "theory": [
                        "Aritmetik Operatörler:",
                        "  + Toplama, - Çıkarma, * Çarpma, / Bölme",
                        "  % Mod (kalan), ** Üs, // Tam bölme",
                        "",
                        "Karşılaştırma Operatörleri:",
                        "  == Eşit, != Eşit değil, > Büyük, < Küçük",
                        "  >= Büyük eşit, <= Küçük eşit",
                        "",
                        "Mantıksal Operatörler:",
                        "  and Ve, or Veya, not Değil"
                    ]
                }
            },
            "otomasyon_egitim": {
                "dosya_okuma": {
                    "title": "Dosya Okuma/Yazma Otomasyonu",
                    "theory": [
                        "Dosya işlemleri için open() fonksiyonu kullanılır:",
                        "Modlar: 'r' okuma, 'w' yazma, 'a' ekleme",
                        "with open() kullanımı dosyayı otomatik kapatır",
                        "",
                        "Önemli Fonksiyonlar:",
                        "  read(): Tüm dosyayı okur",
                        "  readline(): Bir satır okur",
                        "  readlines(): Tüm satırları liste olarak okur",
                        "  write(): Dosyaya yazar",
                        "  close(): Dosyayı kapatır"
                    ]
                }
            }
        }
        
        return lessons_content.get(module_key, {}).get(lesson_key, {"title": "Ders", "theory": ["İçerik hazırlanıyor..."]})

    def get_code_example(self, module_key, lesson_key):
        """Kod örneğini getir"""
        code_examples = {
            "degiskenler": '''
# Değişken tanımlama
isim = "Ahmet"
yas = 25
boy = 1.75
ogrenci = True

# Veri tiplerini yazdırma
print("İsim:", isim, "Tip:", type(isim))
print("Yaş:", yas, "Tip:", type(yas))
print("Boy:", boy, "Tip:", type(boy))
print("Öğrenci:", ogrenci, "Tip:", type(ogrenci))

# Liste ve sözlük
meyveler = ["elma", "armut", "muz"]
kisi = {"ad": "Mehmet", "yas": 30}

print("Meyveler:", meyveler)
print("Kişi:", kisi)
''',
            "operatorler": '''
# Aritmetik operatörler
a = 10
b = 3

print("Toplam:", a + b)
print("Fark:", a - b)
print("Çarpım:", a * b)
print("Bölüm:", a / b)
print("Tam Bölüm:", a // b)
print("Kalan:", a % b)
print("Üs:", a ** b)

# Karşılaştırma operatörleri
x = 5
y = 8

print("x == y:", x == y)
print("x != y:", x != y)
print("x > y:", x > y)
print("x < y:", x < y)

# Mantıksal operatörler
dogru = True
yanlis = False

print("dogru and yanlis:", dogru and yanlis)
print("dogru or yanlis:", dogru or yanlis)
print("not dogru:", not dogru)
''',
            "dosya_okuma": '''
# Dosya yazma
with open("ornek.txt", "w", encoding="utf-8") as dosya:
    dosya.write("Merhaba Dünya!\\n")
    dosya.write("Python ile dosya işlemleri\\n")
    dosya.write("Otomasyon eğitimi\\n")

print("Dosya yazma tamamlandı!")

# Dosya okuma
print("\\nDosya içeriği:")
with open("ornek.txt", "r", encoding="utf-8") as dosya:
    icerik = dosya.read()
    print(icerik)

# Satır satır okuma
print("Satır satır okuma:")
with open("ornek.txt", "r", encoding="utf-8") as dosya:
    satirlar = dosya.readlines()
    for i, satir in enumerate(satirlar, 1):
        print(f"{i}. satır: {satir.strip()}")
'''
        }
        
        return code_examples.get(lesson_key, '# Kod örneği hazırlanıyor...')

    def run_code_example(self, code):
        """Kod örneğini çalıştır"""
        try:
            print(f"\n{Fore.YELLOW}╔══════════ KOD ÇALIŞTIRILIYOR ══════════╗")
            print(f"║{Style.RESET_ALL}")
            exec(code)
            print(f"║{Style.RESET_ALL}")
            print(f"╚══════════════════════════════════════════╝")
        except Exception as e:
            print(f"{Fore.RED}Hata oluştu: {e}")

    def code_examples(self):
        """Kod örneklerini incele"""
        examples = {
            "1": "Web Scraping Örneği",
            "2": "Excel Otomasyonu", 
            "3": "E-posta Gönderme",
            "4": "Dosya Organizatörü",
            "5": "Veritabanı İşlemleri"
        }
        
        while True:
            self.clear_screen()
            self.print_menu("KOD ÖRNEKLERİ", examples)
            
            choice = input(f"\n{Fore.CYAN}Seçiminiz (1-5, 0=Ana menü): {Style.RESET_ALL}")
            
            if choice == "0":
                break
            elif choice in examples:
                self.show_code_example(choice)
            else:
                input(f"{Fore.RED}Geçersiz seçim! (Enter)")

    def show_code_example(self, example_key):
        """Kod örneğini göster"""
        examples = {
            "1": '''
# Web Scraping Örneği
import requests
from bs4 import BeautifulSoup

def basit_scraper(url):
    try:
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Başlıkları al
        basliklar = soup.find_all('h1')[:3]
        print("Sayfa Başlıkları:")
        for baslik in basliklar:
            print(f"- {baslik.text.strip()}")
            
    except Exception as e:
        print(f"Hata: {e}")

# Kullanım
basit_scraper("https://example.com")
''',
            "2": '''
# Excel Otomasyonu
import openpyxl
from openpyxl import Workbook

def excel_olustur():
    # Yeni workbook oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Veriler"
    
    # Başlıklar
    ws['A1'] = 'İsim'
    ws['B1'] = 'Yaş'
    ws['C1'] = 'Şehir'
    
    # Veriler
    veriler = [
        ['Ahmet', 25, 'İstanbul'],
        ['Ayşe', 30, 'Ankara'],
        ['Mehmet', 35, 'İzmir']
    ]
    
    for i, veri in enumerate(veriler, 2):
        ws[f'A{i}'] = veri[0]
        ws[f'B{i}'] = veri[1]
        ws[f'C{i}'] = veri[2]
    
    # Kaydet
    wb.save('ornek_veriler.xlsx')
    print("Excel dosyası oluşturuldu: ornek_veriler.xlsx")

excel_olustur()
'''
        }
        
        self.clear_screen()
        print(f"\n{Fore.CYAN}╔══════════ KOD ÖRNEĞİ ══════════╗")
        print(f"║{Style.RESET_ALL}")
        
        code = examples.get(example_key, "# Örnek hazırlanıyor...")
        for line in code.split('\n'):
            print(f"║ {Fore.GREEN}{line}{Style.RESET_ALL}")
        
        print(f"║{Style.RESET_ALL}")
        print(f"╚══════════════════════════════════╝")
        
        input(f"\n{Fore.CYAN}Devam etmek için Enter...{Style.RESET_ALL}")

    def practice_exercises(self):
        """Alıştırmalar"""
        exercises = {
            "1": "Hesap Makinesi Yapımı",
            "2": "Şifre Üretici",
            "3": "Sayı Tahmin Oyunu"
        }
        
        self.clear_screen()
        self.print_menu("ALIŞTIRMALAR", exercises)
        
        choice = input(f"\n{Fore.CYAN}Seçiminiz (1-3): {Style.RESET_ALL}")
        
        if choice == "1":
            self.calculator_exercise()
        elif choice == "2":
            self.password_generator_exercise()
        else:
            input(f"{Fore.YELLOW}Bu alıştırma yakında eklenecek! (Enter)")

    def calculator_exercise(self):
        """Hesap makinesi alıştırması"""
        print(f"\n{Fore.CYAN}╔══════════ HESAP MAKİNESİ ALIŞTIRMASI ══════════╗")
        print(f"║{Style.RESET_ALL}")
        print(f"║ {Fore.YELLOW}Görev: 4 işlem yapan basit bir hesap makinesi yapın{Style.RESET_ALL}")
        print(f"║ {Fore.GREEN}İpucu: input(), if-elif-else, float() kullanın{Style.RESET_ALL}")
        print(f"║{Style.RESET_ALL}")
        print(f"╚══════════════════════════════════════════════════╝")
        
        input(f"\n{Fore.CYAN}Çözümü görmek için Enter...{Style.RESET_ALL}")
        
        solution = '''
# Hesap Makinesi Çözümü
while True:
    print("\\n--- Hesap Makinesi ---")
    print("1. Toplama")
    print("2. Çıkarma") 
    print("3. Çarpma")
    print("4. Bölme")
    print("5. Çıkış")
    
    secim = input("Seçiminiz (1-5): ")
    
    if secim == '5':
        print("Güle güle!")
        break
    
    if secim in ['1', '2', '3', '4']:
        try:
            sayi1 = float(input("İlk sayı: "))
            sayi2 = float(input("İkinci sayı: "))
            
            if secim == '1':
                sonuc = sayi1 + sayi2
                print(f"Sonuç: {sayi1} + {sayi2} = {sonuc}")
            elif secim == '2':
                sonuc = sayi1 - sayi2
                print(f"Sonuç: {sayi1} - {sayi2} = {sonuc}")
            elif secim == '3':
                sonuc = sayi1 * sayi2
                print(f"Sonuç: {sayi1} × {sayi2} = {sonuc}")
            elif secim == '4':
                if sayi2 != 0:
                    sonuc = sayi1 / sayi2
                    print(f"Sonuç: {sayi1} ÷ {sayi2} = {sonuc}")
                else:
                    print("Hata: Sıfıra bölünemez!")
        except ValueError:
            print("Hata: Geçerli sayı girin!")
    else:
        print("Geçersiz seçim!")
'''
        print(f"\n{Fore.GREEN}Çözüm:{Style.RESET_ALL}")
        for line in solution.split('\n'):
            print(f"{Fore.CYAN}{line}{Style.RESET_ALL}")

    def password_generator_exercise(self):
        """Şifre üretici alıştırması"""
        print(f"\n{Fore.CYAN}╔══════════ ŞİFRE ÜRETİCİ ALIŞTIRMASI ══════════╗")
        print(f"║{Style.RESET_ALL}")
        print(f"║ {Fore.YELLOW}Görev: Rastgele şifre üreten program yapın{Style.RESET_ALL}")
        print(f"║ {Fore.GREEN}İpucu: random modülü, string modülü kullanın{Style.RESET_ALL}")
        print(f"║{Style.RESET_ALL}")
        print(f"╚══════════════════════════════════════════════════╝")
        
        input(f"\n{Fore.CYAN}Çözümü görmek için Enter...{Style.RESET_ALL}")
        
        solution = '''
# Şifre Üretici Çözümü
import random
import string

def sifre_uret(uzunluk=12):
    # Tüm karakterleri birleştir
    tum_karakterler = string.ascii_letters + string.digits + string.punctuation
    
    # Rastgele şifre oluştur
    sifre = ''.join(random.choice(tum_karakterler) for i in range(uzunluk))
    return sifre

# Kullanım
print("Rastgele Şifreler:")
for i in range(5):
    sifre = sifre_uret(10)
    print(f"{i+1}. {sifre}")

# Sadece harf ve rakam
def basit_sifre_uret(uzunluk=8):
    karakterler = string.ascii_letters + string.digits
    return ''.join(random.choice(karakterler) for i in range(uzunluk))

print("\\nBasit Şifreler:")
for i in range(3):
    print(f"{i+1}. {basit_sifre_uret(6)}")
'''
        print(f"\n{Fore.GREEN}Çözüm:{Style.RESET_ALL}")
        for line in solution.split('\n'):
            print(f"{Fore.CYAN}{line}{Style.RESET_ALL}")

    def settings_menu(self):
        """Ayarlar menüsü"""
        while True:
            self.clear_screen()
            settings = {
                "1": "Kullanıcı İstatistikleri",
                "2": "Verileri Sıfırla",
                "0": "Ana Menü"
            }
            
            self.print_menu("AYARLAR", settings)
            
            choice = input(f"\n{Fore.CYAN}Seçiminiz: {Style.RESET_ALL}")
            
            if choice == "1":
                self.show_statistics()
            elif choice == "2":
                self.reset_data()
            elif choice == "0":
                break
            else:
                input(f"{Fore.RED}Geçersiz seçim! (Enter)")

    def show_statistics(self):
        """İstatistikleri göster"""
        self.clear_screen()
        completed = len(self.user_data["completed_lessons"])
        total_lessons = sum(len(module["lessons"]) for module in self.modules.values())
        
        print(f"\n{Fore.CYAN}╔══════════════ İSTATİSTİKLER ══════════════╗")
        print(f"║ {Fore.GREEN}Toplam Ders: {total_lessons}")
        print(f"║ {Fore.BLUE}Tamamlanan: {completed}")
        print(f"║ {Fore.YELLOW}Tamamlanma Oranı: {(completed/total_lessons)*100:.1f}%")
        print(f"║ {Fore.MAGENTA}Toplam Puan: {self.user_data['score']}")
        print(f"║ {Fore.CYAN}Başlangıç Tarihi: {self.user_data['start_date']}")
        
        # Modül bazlı istatistikler
        print(f"║")
        print(f"║ {Fore.WHITE}Modül İlerlemeleri:")
        for module_key, module in self.modules.items():
            mod_completed = sum(1 for lesson in module["lessons"] if lesson in self.user_data["completed_lessons"])
            mod_total = len(module["lessons"])
            print(f"║   {module['title']}: {mod_completed}/{mod_total}")
        
        print(f"╚═══════════════════════════════════════════════╝")
        
        input(f"\n{Fore.CYAN}Devam etmek için Enter...{Style.RESET_ALL}")

    def reset_data(self):
        """Verileri sıfırla"""
        confirm = input(f"\n{Fore.RED}Tüm verileriniz silinecek! Emin misiniz? (e/h): {Style.RESET_ALL}").lower()
        if confirm == 'e':
            self.user_data = {
                "start_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "completed_lessons": [],
                "current_module": "python_temelleri",
                "score": 0
            }
            self.save_user_progress()
            print(f"{Fore.GREEN}✓ Veriler sıfırlandı!")
            time.sleep(2)

def main():
    """Ana fonksiyon"""
    try:
        app = EkomCode()
        app.main_menu()
    except KeyboardInterrupt:
        print(f"\n\n{Fore.YELLOW}Program kapatılıyor...")
        print(f"{Fore.GREEN}EkomCode'u kullandığınız için teşekkürler! 🚀")

if __name__ == "__main__":
    main()
